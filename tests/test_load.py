"""Tests for the load stage."""

from __future__ import annotations

from decimal import Decimal
from pathlib import Path
from unittest.mock import MagicMock, patch

import openpyxl

from invoice_etl.load.db_loader import load_invoice
from invoice_etl.load.excel_loader import load_invoice_to_excel
from invoice_etl.models.invoice import Invoice, LineItem


def _make_invoice(**kwargs: object) -> Invoice:
    defaults: dict[str, object] = {
        "invoice_number": "INV-TEST-001",
        "total_amount": Decimal("100.00"),
    }
    defaults.update(kwargs)
    return Invoice(**defaults)  # type: ignore[arg-type]


def _make_mock_engine(invoice_id: int | None = 42) -> tuple[MagicMock, MagicMock]:
    """Return a (mock_engine, mock_conn) pair configured with the given invoice id.

    When ``invoice_id`` is ``None``, simulates an ON CONFLICT DO NOTHING result.
    """
    mock_engine = MagicMock()
    mock_conn = MagicMock()
    mock_result = MagicMock()
    mock_result.fetchone.return_value = (invoice_id,) if invoice_id is not None else None
    mock_conn.execute.return_value = mock_result
    mock_engine.begin.return_value.__enter__.return_value = mock_conn
    return mock_engine, mock_conn


def test_load_returns_invoice_id() -> None:
    mock_engine, _ = _make_mock_engine(invoice_id=42)
    invoice_id = load_invoice(_make_invoice(), engine=mock_engine)
    assert invoice_id == 42


def test_load_returns_minus_one_on_duplicate() -> None:
    mock_engine, _ = _make_mock_engine(invoice_id=None)
    invoice_id = load_invoice(_make_invoice(), engine=mock_engine)
    assert invoice_id == -1


def test_load_passes_customer_number_to_invoices_insert() -> None:
    """load_invoice forwards customer_number to the invoices INSERT parameter dict."""
    mock_engine, mock_conn = _make_mock_engine(invoice_id=42)
    invoice = _make_invoice(customer_number="081997")
    load_invoice(invoice, engine=mock_engine)
    # The first execute call is the invoices INSERT; its parameter dict is args[1].
    assert mock_conn.execute.call_args_list[0].args[1]["customer_number"] == "081997"


def test_load_passes_new_line_item_fields_to_insert() -> None:
    """load_invoice forwards the four new LineItem fields to the line_items INSERT."""
    mock_engine, mock_conn = _make_mock_engine(invoice_id=42)
    line_item = LineItem(
        item="01553",
        store_number="00562",
        offer_number="545039",
        unit_of_measure="EA",
    )
    invoice = _make_invoice(line_items=[line_item])
    load_invoice(invoice, engine=mock_engine)
    # The second execute call is the line_items INSERT; verify the parameter dict.
    line_item_params: dict[str, object] = mock_conn.execute.call_args_list[1].args[1]
    assert line_item_params["item"] == "01553"
    assert line_item_params["store_number"] == "00562"
    assert line_item_params["offer_number"] == "545039"
    assert line_item_params["unit_of_measure"] == "EA"


# ---------------------------------------------------------------------------
# excel_loader tests
# ---------------------------------------------------------------------------


def _make_captured_workbook() -> tuple[openpyxl.Workbook, MagicMock]:
    """Return a (real_wb, mock_save) pair for filesystem-free excel_loader tests.

    The real workbook is used so sheet names and cell contents can be inspected
    after the function under test writes to it.  The save method is replaced with
    a no-op mock so no file is written to disk.

    Returns:
        A tuple of (real Workbook instance, MagicMock replacing ``save``).
    """
    real_wb = openpyxl.Workbook()
    mock_save = MagicMock()
    real_wb.save = mock_save  # replacing save to prevent filesystem write in tests
    return real_wb, mock_save


def test_load_invoice_to_excel_returns_output_path() -> None:
    """load_invoice_to_excel returns the output_path passed as argument."""
    invoice = Invoice(invoice_number="INV-001")
    real_wb, _ = _make_captured_workbook()

    with patch.object(openpyxl, "Workbook", return_value=real_wb):
        result = load_invoice_to_excel(invoice, Path("out/invoice.xlsx"))

    assert result == Path("out/invoice.xlsx")


def test_load_invoice_to_excel_creates_invoice_and_line_items_sheets() -> None:
    """load_invoice_to_excel produces exactly two sheets: Invoice and LineItems."""
    invoice = Invoice(invoice_number="INV-001")
    real_wb, _ = _make_captured_workbook()

    with patch.object(openpyxl, "Workbook", return_value=real_wb):
        load_invoice_to_excel(invoice, Path("invoice.xlsx"))

    assert real_wb.sheetnames == ["Invoice", "LineItems"]


def test_load_invoice_to_excel_invoice_sheet_has_correct_column_headers() -> None:
    """The Invoice sheet header row contains all expected invoice field names."""
    invoice = Invoice(invoice_number="INV-001")
    real_wb, _ = _make_captured_workbook()

    with patch.object(openpyxl, "Workbook", return_value=real_wb):
        load_invoice_to_excel(invoice, Path("invoice.xlsx"))

    ws = real_wb["Invoice"]
    # Header row is row 1; verify the first and last expected column names.
    assert ws.cell(row=1, column=1).value == "invoice_number"
    assert ws.cell(row=1, column=13).value == "source_file"


def test_load_invoice_to_excel_line_items_sheet_has_correct_column_headers() -> None:
    """The LineItems sheet header row contains all expected line-item field names."""
    invoice = Invoice(invoice_number="INV-001")
    real_wb, _ = _make_captured_workbook()

    with patch.object(openpyxl, "Workbook", return_value=real_wb):
        load_invoice_to_excel(invoice, Path("invoice.xlsx"))

    ws = real_wb["LineItems"]
    assert ws.cell(row=1, column=1).value == "description"
    assert ws.cell(row=1, column=9).value == "unit_of_measure"


def test_load_invoice_to_excel_writes_one_data_row_per_line_item() -> None:
    """The LineItems sheet contains one data row for each line item in the invoice."""
    line_items = [
        LineItem(description="Widget A", quantity=Decimal("2"), unit_price=Decimal("5.00")),
        LineItem(description="Widget B", quantity=Decimal("3"), unit_price=Decimal("10.00")),
    ]
    invoice = Invoice(invoice_number="INV-002", line_items=line_items)
    real_wb, _ = _make_captured_workbook()

    with patch.object(openpyxl, "Workbook", return_value=real_wb):
        load_invoice_to_excel(invoice, Path("invoice.xlsx"))

    ws = real_wb["LineItems"]
    # Row 1 is the header; rows 2 and 3 are the two line items.
    assert ws.cell(row=2, column=1).value == "Widget A"
    assert ws.cell(row=3, column=1).value == "Widget B"


def test_load_invoice_to_excel_saves_to_output_path() -> None:
    """load_invoice_to_excel calls save with the supplied output path."""
    invoice = Invoice(invoice_number="INV-001")
    real_wb, mock_save = _make_captured_workbook()
    output_path = Path("exports/invoice.xlsx")

    with patch.object(openpyxl, "Workbook", return_value=real_wb):
        load_invoice_to_excel(invoice, output_path)

    mock_save.assert_called_once_with(output_path)
